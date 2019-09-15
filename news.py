import re
import os
import requests
from openpyxl import Workbook 
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from retry import retry
from bs4 import BeautifulSoup
from aip import AipNlp
from pprint import pprint

APP_ID = '17237009'
API_KEY = 'i1RVfu7Bazbf6zD22cbERKP0'
SECRET_KEY = 'ccesNDkICGhj5lLkNbdvzdobnMR9pDbP'

client = AipNlp(APP_ID, API_KEY, SECRET_KEY)

origin_path = os.getcwd()

@retry(tries = 5, delay = 2)
def getcontent(url):
    response = requests.get(url)
    web = BeautifulSoup(response.text,"html.parser")
    #pagetype = web.find_all("meta"）
    for a in web.find_all("meta",rel = True):
        if "roll" in str(a["href"]): 
            pagetype = 1
            articletype = "快讯"
        elif "morning" in str(a["href"]): 
            pagetype = 2
            articletype = "内参"
        else:
            pagetype = 3
            articletype = "深度"
        #print(pagetype)
        #print(articletype)

    '''    
    if pagetype:
        heads = web.find_all("div",attrs = {"class":"jsx-1016208558 title fontBold"})
    else: 
        heads = web.find_all("div",attrs = {"class":"jsx-1016208558 title fontBold"})
    for a in heads:
        title = a.string
        print(title)
    '''
    if pagetype != 1:
        title = web.title.string
        #title.replace("\xa0"," ")
        #title.encode("gbk","ignore").decode("gbk","replace")
        sent_title_origin = client.sentimentClassify(title)
        sent_title = sent_title_origin["items"][0]
    #keyword = client.keyword(title)    

    else:
        title = "电报：无标题"
        sent_title = {'positive_prob': 0, 'confidence': 0, 'negative_prob': 0, 'sentiment': 0}
        #print(title)

    '''
    if pagetype:
        depths = web.find_all("a", attrs = {"href" : "/depth"})
    else:
        depths = web.find_all("a", attrs = {"href" : "/telegraph"})
        
    for a in depths:
        depth = a.string
        print(depth)
        break
    '''

    times = web.find_all("div",attrs = {"class":"jsx-1016208558 ctime"})
    for a in times:
        dateandtime = a.string
        date = dateandtime[0:10]
        time = dateandtime[-8:]
        #print(date,time)

    if pagetype == 3:
        writers = web.find_all("span",attrs = {"class":"jsx-1016208558 writer"})
        for a in writers:
            #print(a)
            if "|" not in a.string:
                writer = "/"
                media = a.string
            else:
                writerandfirm = a.string.split("|")
                writer = writerandfirm[0]
                media = writerandfirm[1]
            #print(writer,media)
    else:
        writer = "／"
        media = "财联社"
        #print(writer,media)


    reads = web.find_all("div",attrs = {"class":"jsx-1016208558 readNum"})
    for a in reads:
        a= str(a)
        take = re.compile("-->"+'(.*?)'+"</",re.S)
        b = take.findall(a) #取字符串中的阅读数量，findall结果为列表
        readnum = int(b[0]) #取列表的数字
        #print(readnum)

    if pagetype != 1:
        origins = web.find_all("span",attrs = {"class":"jsx-1016208558 tag"})
        if origins == []:
            origin = "/"
        else:    
            for a in origins:
                origin = a.string
                #print(origin)
    else:
        origin = "原创"
        #print(origin)

    Bigcontents = web.find_all("div",attrs = {"class":"jsx-1016208558 thisContent c-000"})
    content = ""
    writer_k = 1
    writer_list = []
    for a in Bigcontents:
        if pagetype == 1:
            content = a.string     
        else:
            ptag = a.find_all("p")
            for b in ptag:
                c = b.string
                if c == None :
                    c = "/n"
                getwrtier_list = ["作者","本报记者","特约作者","记者"]
                if writer_k == 1:
                    for i in getwrtier_list:
                        if i in c:
                            d = re.compile(i+' (.*?)' + "/n",re.S).findall(c)
                            e = re.compile(i+' (.*?)' + "报道",re.S).findall(c)
                            f = re.compile(i+' (.*?)' + "  ",re.S).findall(c)
                            g = re.compile(i+' (.*?)' + "）",re.S).findall(c)
                            h = re.findall("(?<=" +i + ")(.*?)", c)
                            print(d,e,f,g,h)
                            #print(len(d),len(e),len(f),len(g),len(h))
                            if  d:
                                writer_list.append(d[0])
                            if  e:
                                writer_list.append(e[0])
                            if  f :
                                writer_list.append(f[0])
                            if  g :
                                writer_list.append(g[0])
                            #if  h :
                            #    writer_list.append(h[0])
                            if not writer_list:
                                writer = c
                               # print(writer)
                            #print(writer_list)
                            else :
                                writer = min(writer_list,key = len)
                            writer.replace(" ","") 
                                #print(writer)
                            break                        
                writer_k += 1
                content = content + c

        #keyword = client.keyword(content)
        content.replace(u'\xa0',u"")
        content.replace(" ","")
        error_report = 0
        content = ''.join(content.split())
        n = len(content)//1024
        #print(len(content)/1024)
        if len(content)/1024 > 1:
            content = content[0:1023]
        try:
            sent_content = client.sentimentClassify(content)
        except:
            print("错误")
            return
        if "items" in sent_content.keys():
            sent_content_items = sent_content["items"][0]
            [a,b,c,d]=list(sent_content_items.values())
        else:
            [a,b,c,d] = ["","","",""]
        '''
        if (len(content)/1024) >1: #如果文本太大，分割加权
            #temp_sent_content_dict = {'positive_prob':0,'cofidence':0,'negative_prob':0,'sentiment':0}
            sent_content_list=[0,0,0,0]
            for i in range(0,n+1):
                if ((i+1)*1024) >len(content):
                    temp_content = content[i*1024:]
                    #print(temp_content)
                    #print("last time")
                    try:
                        temp_sent_content = client.sentimentClassify(temp_content)
                        print("最后一次，能够获得")
                        #print(temp_sent_content["items"][0])
                    except: 
                        list_content = []
                        error_report += 1 
                        print("调用错误：\xa0；错误次数" + str(error_report))
                        return list_content                        
                else:
                    temp_content = content[i*1024:(i+1)*1024-1]
                    try:
                        temp_sent_content = client.sentimentClassify(temp_content)
                        print("其中一次，能够获得")
                    except: 
                        list_content = []
                        error_report += 1 
                        print("调用错误：\xa0；错误次数" + str(error_report))
                        return list_content 
                
                temp_dict = temp_sent_content["items"][0] 
                temp_sent_content_list = list(temp_dict.values())
                #取到items里面的字典
                sent_content_list = [sent_content_list[i]+temp_sent_content_list[i] for i in range(0,3)]
                #[a,b,c,d] = sent_content_list
                print(sent_content_list)
                #print(temp_dict,len(temp_content),len(content)) #字典相加
                #temp_sent_content_dict['positive_prob'] += temp_dict['positive_prob']*len(temp_content)/len(content)
                #temp_sent_content_dict['cofidence'] += temp_dict['confidence']*len(temp_content)/len(content)
                #temp_sent_content_dict['negative_prob'] += temp_dict['negative_prob']*len(temp_content)/len(content)
                #temp_sent_content_dict['sentiment'] += temp_dict['sentiment']*len(temp_content)/len(content)
            #sent_content = temp_sent_content_dict
            #print(sent_content)
            
        else:
            try:
                sent_content = client.sentimentClassify(content)
            except:
                list_content = []
                error_report += 1 
                print("调用错误：\xa0；错误次数" + str(error_report))
                return list_content
            sent_content = sent_content["items"][0]
            [a,b,c,d] = list(sent_content.values())
        '''

    list_content = [url,articletype, title, date, time, writer, media, origin, readnum, content, sent_title["positive_prob"], sent_title["confidence"], sent_title["negative_prob"], sent_title["sentiment"], a,b,c,d]
    
    code_list = find_stkcd(content) #匹配企业名称
    macro = 0
    market = 0
    if code_list :
        list_content.append(macro)
        list_content.append(market)
        for related_code in code_list:
            list_content.append(related_code)
    elif find_macro(content) == 1:  #匹配宏观
        macro = find_macro(content)
        list_content.append(macro)
        list_content.append(market)
    else : #匹配市场
        list_content.append(macro)
        list_content.append(market)   
    
    print(list_content)
    return list_content
    #return articletype, title, date, time, writer, media, readnum, content

def savexlsx(r,list_content):
    wb = load_workbook("财联社新闻汇总.xlsx")
    sheet = wb["Sheet"]
    c = 1
    for item in list_content:
        sheet.cell(row = r, column = c).value = item
        c += 1
    
    wb.save("财联社新闻汇总.xlsx")

def find_stkcd(content): #寻找匹配的企业名字
    data = pd.read_csv("codenum.csv",usecols = [0],encoding="utf-8")
    name = pd.read_csv("codenum.csv",usecols=[1],encoding="utf-8")
    #print(data)
    code_list_numpy = np.array(data)
    code_list_origin = code_list_numpy.tolist()
    name_list_nummpy = np.array(name)
    name_list_origin = name_list_nummpy.tolist()
    #print(type(code_list))
    code_list_new = []
    for itemlist in code_list_origin:
        for item in itemlist:
            code_list_new.append(item)
    code_list = []
    name_list = []
    for item in code_list_new:
        code = "%06d" %item
        code = str(code)
        if code in content: #content 在list_content 中的位置为第10列  
            code_list.append(code) 
    for namelist in name_list_origin:
        for name in namelist:
            name_list.append(name)
            if name in content:
                p = name_list.index(name)
                code_list.append(code_list_new[p])
    return code_list

def find_macro(content): #判断是否宏观信息
    macro = 0
    macro_data = pd.read_csv("宏观经济新闻词汇.csv",encoding="utf-8")
    macro_word_list_nummy = np.array(macro_data)
    macro_word_lsit_origin = macro_word_list_nummy.tolist()
    for itemlist in macro_word_lsit_origin:
        for item in itemlist:
            if item in content:
                macro = 1
    return macro

def find_market_news(content): #判断是否市场行情信息
    market = 0
    market_data =["沪深两市","创业板","A股","深成指"]
    for item in market_data:
        if item in content:
            market = 1
    return market

def main():
    url_head = "https://www.cls.cn/depth/"
    wb = Workbook() 
    sheet = wb.active
    label_list =["url","articletype", "title", "date", "time", "writer", "media", "origin", "readnum", "content", "sent_title_positive_prob", "sent_title_confidence", "sent_title_negative_prob", "sent_title_negative_prob", "sent_title_sentiment", "sent_content_positive_prob","sent_content_confidence","sent_content_negative_prob","sent_content_sentiment","macro","market","code"]
    c = 1
    for item in label_list:
        sheet.cell(row = 1, column = c).value = item 
        c += 1
    wb.save("财联社新闻汇总.xlsx")

    r = 2
    for i in range(229,400) : 
        url_tail = "%06d" %i #补齐前面的0
        url_tail = str(url_tail)
        url = url_head + "000330"    
        list_content = getcontent(url)
        print(list_content)
        
        if list_content == []:
            continue
        elif list_content ==None:
            continue
        else:
            savexlsx(r,list_content)
            r += 1
        

if __name__ == "__main__":
    main()
